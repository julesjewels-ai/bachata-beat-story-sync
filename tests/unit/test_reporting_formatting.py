"""
Unit tests for report formatting logic.
"""
import openpyxl
from openpyxl.formatting.rule import Rule
from src.services.reporting.formatting import ReportFormatter


def test_apply_header_style():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    cell = ws.cell(row=1, column=1, value="Header")

    formatter = ReportFormatter()
    formatter.apply_header_style(cell, center=True)

    assert cell.font.bold is True
    assert cell.alignment.horizontal == "center"


def test_adjust_column_widths():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(row=1, column=1, value="Short")
    ws.cell(row=2, column=1, value="A very long string that should determine width")

    formatter = ReportFormatter()
    formatter.adjust_column_widths(ws)

    col_dim = ws.column_dimensions["A"]
    # Length of long string is 44. Width should be 46.
    assert col_dim.width >= 44
    assert col_dim.width <= 52  # Cap is 50 + 2


def test_apply_intensity_conditional_formatting():
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    formatter = ReportFormatter()

    # Apply to Column C (3), Rows 2 to 5
    formatter.apply_intensity_conditional_formatting(ws, 2, 5, 3)

    # Verify rule addition
    found_rule = None
    for cf in ws.conditional_formatting:
        if "C2:C5" in str(cf.sqref):
            if cf.rules:
                found_rule = cf.rules[0]
            break

    assert found_rule is not None, "Conditional formatting rule not found for C2:C5"

    # In openpyxl, ColorScaleRule factory returns a Rule object
    assert isinstance(found_rule, Rule)
    # Check if it has a colorScale attribute (it should for this type)
    assert hasattr(found_rule, 'colorScale')

    # Verify colors
    cs = found_rule.colorScale
    assert cs is not None
    # cfvo (Conditional Formatting Value Objects) should be 3 (min, mid, max)
    assert len(cs.cfvo) == 3
    # Colors should be 3
    assert len(cs.color) == 3

    assert cs.color[0].rgb == "00F8696B"
    assert cs.color[2].rgb == "0063BE7B"
