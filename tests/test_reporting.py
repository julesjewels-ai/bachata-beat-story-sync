"""
Tests for the ExcelReportGenerator.
"""
import os
import openpyxl
from src.services.reporting.generator import ExcelReportGenerator
from src.core.models import AudioAnalysisResult, VideoAnalysisResult, AudioSection


def test_generate_report(tmp_path):
    output_path = tmp_path / "test_report.xlsx"

    audio_data = AudioAnalysisResult(
        filename="bachata.wav",
        bpm=128.0,
        duration=180.0,
        peaks=[10.5, 20.0],
        sections=[
            AudioSection(start_time=0.0, end_time=30.0, duration=30.0, label="Intro"),
            AudioSection(start_time=30.0, end_time=60.0, duration=30.0, label="Verse")
        ],
        beat_times=[0.5, 1.0],
        intensity_curve=[0.5, 0.6]
    )

    video_data = [
        VideoAnalysisResult(
            path="/videos/clip1.mp4",
            intensity_score=0.8,
            duration=10.0,
            thumbnail_data=None
        ),
        VideoAnalysisResult(
            path="/videos/clip2.mp4",
            intensity_score=0.4,
            duration=15.0,
            thumbnail_data=None
        )
    ]

    generator = ExcelReportGenerator()
    result = generator.generate_report(
        audio_data, video_data, str(output_path)
    )

    assert result == str(output_path)
    assert os.path.exists(output_path)

    # Verify Content
    wb = openpyxl.load_workbook(output_path)
    assert "Analysis Summary" in wb.sheetnames
    assert "Video Library" in wb.sheetnames

    ws_summary = wb["Analysis Summary"]
    assert ws_summary["B2"].value == "bachata.wav"
    assert ws_summary["B3"].value == 128.0

    # Verify sections string format
    # Row 6 is sections
    sections_cell = ws_summary["B6"]
    assert "Intro (0.0-30.0s)" in sections_cell.value
    assert "Verse (30.0-60.0s)" in sections_cell.value

    ws_videos = wb["Video Library"]
    assert ws_videos["A2"].value == "/videos/clip1.mp4"
    assert ws_videos["C2"].value == 0.8
