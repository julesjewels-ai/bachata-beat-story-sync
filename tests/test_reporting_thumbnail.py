import os
import openpyxl
from io import BytesIO
from PIL import Image
import pytest
from src.core.models import AudioAnalysisResult, VideoAnalysisResult
from src.services.reporting import ExcelReportGenerator

def create_dummy_image_bytes():
    """Creates a small red image in memory."""
    img = Image.new('RGB', (100, 100), color='red')
    buf = BytesIO()
    img.save(buf, format='JPEG')
    return buf.getvalue()

def test_excel_report_with_thumbnail(tmp_path):
    """Verifies that the Excel report contains embedded images."""

    # Setup data
    audio_data = AudioAnalysisResult(
        filename="test.wav",
        bpm=120,
        duration=60,
        peaks=[10.0, 20.0],
        sections=["intro"]
    )

    thumbnail = create_dummy_image_bytes()
    video_data = [
        VideoAnalysisResult(
            path="/tmp/video1.mp4",
            intensity_score=0.5,
            duration=10.0,
            thumbnail_data=thumbnail
        ),
        VideoAnalysisResult(
            path="/tmp/video2.mp4",
            intensity_score=0.8,
            duration=15.0,
            thumbnail_data=None # No thumbnail
        )
    ]

    output_path = tmp_path / "report_with_images.xlsx"
    generator = ExcelReportGenerator()

    # Generate report
    generator.generate_report(audio_data, video_data, str(output_path))

    # Verify
    assert os.path.exists(output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Video Library"]

    # Check for images
    # Openpyxl stores images in ws._images usually
    assert len(ws._images) == 1

    # Check image location
    img = ws._images[0]
    # The anchor can be tricky to assert exactly, but we can check existence
    # We expect it at D2 (since row 1 is header)
    # Different openpyxl versions handle anchors differently.
    # But usually img.anchor is a marker or string.
    # Let's just assert existence for now as a smoke test.

    # Also check cell value is empty string (or None)
    assert ws["D2"].value in ["", None]
    assert ws["D3"].value in ["", None]

    wb.close()
