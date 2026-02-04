"""
Reporting service for generating analysis reports.
"""
import logging
from typing import List, Any
import io
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

from src.core.models import AudioAnalysisResult, VideoAnalysisResult

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Generates Excel reports from analysis data.
    """

    def generate_report(self,
                        audio_data: AudioAnalysisResult,
                        video_data: List[VideoAnalysisResult],
                        output_path: str) -> str:
        """
        Creates an Excel file with analysis details.

        Args:
            audio_data: The audio analysis result.
            video_data: List of video analysis results.
            output_path: Destination path for the .xlsx file.

        Returns:
            The path to the generated file.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active
        if ws_summary:
            ws_summary.title = "Analysis Summary"
            self._write_summary(ws_summary, audio_data, len(video_data))
        else:
            ws_summary = wb.create_sheet(title="Analysis Summary")
            self._write_summary(ws_summary, audio_data, len(video_data))

        # Sheet 2: Video Details
        ws_videos = wb.create_sheet(title="Video Library")
        self._write_video_details(ws_videos, video_data)

        # Sheet 3: Visualizations
        self._add_visualizations(wb, "Video Library", len(video_data))

        wb.save(output_path)
        logger.info(f"Report generated at: {output_path}")
        return output_path

    def _write_table(self, ws, headers: List[str], data: List[Any],
                     bold_first_col: bool = False,
                     center_headers: bool = False):
        """Helper to write standardized tables."""
        self._write_headers(ws, headers, center_headers)
        self._write_rows(ws, data, bold_first_col)
        self._adjust_column_widths(ws)

    def _write_headers(self, ws, headers: List[str], center_headers: bool):
        """Writes table headers to the first row."""
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            if center_headers:
                cell.alignment = Alignment(horizontal="center")

    def _write_rows(self, ws, data: List[Any], bold_first_col: bool):
        """Writes table data starting from the second row."""
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if col_num == 1 and bold_first_col:
                    cell.font = Font(bold=True)

    def _write_summary(self, ws, audio_data: AudioAnalysisResult,
                       video_count: int):
        """Writes summary data to the worksheet."""
        headers = ["Metric", "Value"]
        data = [
            ("Audio File", audio_data.filename),
            ("BPM", audio_data.bpm),
            ("Duration (s)", audio_data.duration),
            ("Peak Count", len(audio_data.peaks)),
            ("Sections", ", ".join(audio_data.sections)),
            ("Total Videos Scanned", video_count)
        ]
        self._write_table(ws, headers, data, bold_first_col=True,
                          center_headers=True)

    def _write_video_details(self, ws, video_data: List[VideoAnalysisResult]):
        """Writes detailed video analysis data."""
        headers = ["Thumbnail", "File Path", "Duration (s)", "Intensity Score"]

        self._write_headers(ws, headers, center_headers=False)

        # Set default width for thumbnail column
        ws.column_dimensions["A"].width = 22  # Approx 160px width

        for row_idx, v in enumerate(video_data, 2):
            # Column 1: Thumbnail
            if v.thumbnail:
                try:
                    img_stream = io.BytesIO(v.thumbnail)
                    pil_img = PILImage.open(img_stream)
                    img = OpenpyxlImage(pil_img)

                    # Anchor image to cell
                    ws.add_image(img, f"A{row_idx}")

                    # Adjust row height to fit image
                    # Points ~ Pixels * 0.75
                    ws.row_dimensions[row_idx].height = pil_img.height * 0.75
                except Exception as e:
                    logger.warning(
                        f"Failed to embed thumbnail for {v.path}: {e}"
                    )
                    ws.cell(row=row_idx, column=1, value="[Error]")
            else:
                ws.cell(row=row_idx, column=1, value="")

            # Column 2: File Path
            ws.cell(row=row_idx, column=2, value=v.path)

            # Column 3: Duration
            ws.cell(row=row_idx, column=3, value=v.duration)

            # Column 4: Intensity Score
            ws.cell(row=row_idx, column=4, value=v.intensity_score)

        # Auto adjust columns B, C, D
        for col_idx in range(2, 5):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            # Header
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val:
                max_len = len(str(header_val))

            # Data
            for row in range(2, len(video_data) + 2):
                val = ws.cell(row=row, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))

            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def _adjust_column_widths(self, ws):
        """Auto-adjusts column widths based on content."""
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            length = min(length, 50)  # Cap width
            col_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[col_letter].width = length + 2

    def _add_visualizations(self, wb, video_sheet_name: str, data_count: int):
        """
        Adds a chart to visualize video intensity.

        Args:
            wb: The openpyxl Workbook object.
            video_sheet_name: Name of the sheet containing video data.
            data_count: Number of video entries.
        """
        if data_count == 0:
            return

        ws = wb.create_sheet(title="Visualizations")

        # Create Chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Video Intensity Distribution"
        chart.y_axis.title = "Intensity Score"
        chart.x_axis.title = "Video Clip Index"

        # Reference Data from Video Library sheet
        # Column 4 is Intensity Score. Row 1 is header.
        data = Reference(
            wb[video_sheet_name],
            min_col=4,
            min_row=1,
            max_row=data_count + 1,
            max_col=4
        )
        chart.add_data(data, titles_from_data=True)

        ws.add_chart(chart, "A1")
