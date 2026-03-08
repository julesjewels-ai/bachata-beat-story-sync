"""
Formatting service for Excel reports.
Handles styles, column sizing, and conditional formatting.
"""

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class ReportFormatter:
    """
    Encapsulates logic for styling Excel worksheets.
    """

    from typing import Any

    def apply_header_style(self, cell: Any, center: bool = False) -> None:
        """Applies header styling (bold) to a cell."""
        cell.font = Font(bold=True)
        if center:
            cell.alignment = Alignment(horizontal="center")

    def apply_bold_font(self, cell: Any) -> None:
        """Makes the cell text bold."""
        cell.font = Font(bold=True)

    def adjust_column_widths(self, ws: Worksheet) -> None:
        """
        Auto-adjusts column widths based on content length.
        Caps width at 50 characters.
        """
        for column_cells in ws.columns:
            # Calculate max length in this column
            length = 0
            for cell in column_cells:
                if cell.value:
                    length = max(length, len(str(cell.value)))

            length = min(length, 50)  # Cap width

            # Add some padding
            adjusted_width = length + 2

            col_idx = column_cells[0].column
            # Ensure column index is valid for mypy
            if not isinstance(col_idx, int):
                continue

            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = adjusted_width

    def apply_intensity_conditional_formatting(
        self, ws: Worksheet, min_row: int, max_row: int, col_idx: int
    ) -> None:
        """
        Applies a 3-color scale to the specified column range.
        Red (Low Intensity) -> Yellow -> Green (High Intensity).

        Args:
            ws: The worksheet to apply formatting to.
            min_row: The starting row number.
            max_row: The ending row number.
            col_idx: The column index (1-based) containing intensity scores.
        """
        if max_row < min_row:
            return

        col_letter = get_column_letter(col_idx)
        range_string = f"{col_letter}{min_row}:{col_letter}{max_row}"

        # Color Scale: Red (Low) -> Yellow (Mid) -> Green (High)
        # F8696B: Light Red
        # FFEB84: Light Yellow
        # 63BE7B: Light Green
        rule = ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(range_string, rule)
