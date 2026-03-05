"""
Main generator service for Excel reports.
Orchestrates the report creation using specialized components.
"""

import logging

import openpyxl

from src.core.models import AudioAnalysisResult, VideoAnalysisResult

from .components import ChartBuilder, ThumbnailEmbedder
from .formatting import ReportFormatter

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Generates Excel reports from analysis data.
    Uses clean architecture to separate formatting and component logic.
    """

    def __init__(self) -> None:
        self.formatter = ReportFormatter()
        self.chart_builder = ChartBuilder()
        self.thumbnail_embedder = ThumbnailEmbedder()

    def generate_report(
        self,
        audio_data: AudioAnalysisResult,
        video_data: list[VideoAnalysisResult],
        output_path: str,
    ) -> str:
        """
        Creates an Excel file with analysis details.

        Args:
            audio_data: The audio analysis result.
            video_data: List of video analysis results.
            output_path: Destination path for the .xlsx file.

        Returns:
            The path to the generated file.
        """
        wb = openpyxl.Workbook()

        # Sheet 1: Summary
        ws_summary = wb.active or wb.create_sheet()
        ws_summary.title = "Analysis Summary"
        self._build_summary_sheet(ws_summary, audio_data, len(video_data))

        # Sheet 2: Video Details
        ws_videos = wb.create_sheet(title="Video Library")
        self._build_video_sheet(ws_videos, video_data)

        # Sheet 3: Visualizations
        self._build_visualization_sheet(wb, "Video Library", len(video_data))

        wb.save(output_path)
        logger.info("Report generated at: %s", output_path)
        return output_path

    from typing import Any

    def _write_headers(self, ws: Any, headers: list[str], center: bool = False) -> None:
        """Helper to write and format table headers."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self.formatter.apply_header_style(cell, center=center)

    def _build_summary_sheet(
        self, ws: Any, audio_data: AudioAnalysisResult, video_count: int
    ) -> None:
        """Constructs the summary sheet."""
        headers = ["Metric", "Value"]
        data = [
            ("Audio File", audio_data.filename),
            ("BPM", audio_data.bpm),
            ("Duration (s)", audio_data.duration),
            ("Peak Count", len(audio_data.peaks)),
            ("Sections", ", ".join(s.label for s in audio_data.sections)),
            ("Total Videos Scanned", video_count),
        ]

        self._write_headers(ws, headers, center=True)

        # Write Data
        for r, (metric, val) in enumerate(data, 2):
            # Metric Name
            c1 = ws.cell(row=r, column=1, value=metric)
            self.formatter.apply_bold_font(c1)
            # Value
            ws.cell(row=r, column=2, value=val)

        self.formatter.adjust_column_widths(ws)

    def _build_video_sheet(self, ws: Any, video_data: list[VideoAnalysisResult]) -> None:
        """Constructs the video details sheet."""
        headers = ["File Path", "Duration (s)", "Intensity Score", "Thumbnail"]

        self._write_headers(ws, headers)

        # Write Rows
        for r, video in enumerate(video_data, 2):
            ws.cell(row=r, column=1, value=video.path)
            ws.cell(row=r, column=2, value=video.duration)
            ws.cell(row=r, column=3, value=video.intensity_score)

            # Embed Thumbnail
            if not video.thumbnail_data:
                ws.cell(row=r, column=4, value="[No Image]")
                continue

            if not self.thumbnail_embedder.embed_thumbnail(
                ws, r, 4, video.thumbnail_data
            ):
                ws.cell(row=r, column=4, value="[Error]")

        # Auto-size columns
        self.formatter.adjust_column_widths(ws)

        # Apply Conditional Formatting to Intensity Score (Column 3)
        row_count = len(video_data)
        if row_count > 0:
            self.formatter.apply_intensity_conditional_formatting(
                ws, min_row=2, max_row=row_count + 1, col_idx=3
            )

    def _build_visualization_sheet(
        self, wb: Any, source_sheet_name: str, data_count: int
    ) -> None:
        """Adds charts to the workbook."""
        if data_count == 0:
            return

        ws = wb.create_sheet(title="Visualizations")
        source_ws = wb[source_sheet_name]

        # Create Intensity Chart (Column 3 is Intensity)
        chart = self.chart_builder.create_intensity_chart(
            source_ws, data_count, intensity_col_idx=3
        )

        if chart:
            ws.add_chart(chart, "A1")
